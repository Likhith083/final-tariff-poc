"""
Conversation Export Service for ATLAS Enterprise
Export AI chat conversations in multiple formats (PDF, JSON, Markdown, Excel).
"""

import asyncio
import json
import uuid
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Union
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
import tempfile
import zipfile

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import markdown
from jinja2 import Template

from ..core.database import get_cache
from ..core.logging import get_logger, log_business_event
from ..core.config import settings
from .enhanced_ai_service import enhanced_ai_service, ConversationMessage, ConversationRole

logger = get_logger(__name__)


class ExportFormat(Enum):
    """Supported export formats."""
    JSON = "json"
    PDF = "pdf"
    MARKDOWN = "markdown"
    HTML = "html"
    EXCEL = "excel"
    TXT = "txt"
    CSV = "csv"


class ExportTemplate(Enum):
    """Export templates."""
    STANDARD = "standard"
    PROFESSIONAL = "professional"
    DETAILED = "detailed"
    SUMMARY = "summary"


@dataclass
class ExportOptions:
    """Export configuration options."""
    format: ExportFormat
    template: ExportTemplate = ExportTemplate.STANDARD
    include_metadata: bool = True
    include_timestamps: bool = True
    include_system_messages: bool = False
    include_knowledge_updates: bool = True
    anonymize_user_data: bool = False
    custom_title: Optional[str] = None
    custom_header: Optional[str] = None
    custom_footer: Optional[str] = None
    page_size: str = "letter"  # letter, a4
    font_size: int = 12


@dataclass
class ExportResult:
    """Export operation result."""
    success: bool
    file_path: Optional[str] = None
    file_name: Optional[str] = None
    file_size: Optional[int] = None
    export_id: Optional[str] = None
    download_url: Optional[str] = None
    error: Optional[str] = None
    metadata: Dict[str, Any] = None


class ConversationExportService:
    """Service for exporting AI chat conversations."""
    
    def __init__(self):
        """Initialize conversation export service."""
        self.cache = None
        self.export_dir = Path("data/exports")
        self.temp_dir = Path("data/temp")
        self._initialized = False
        
        # HTML template for conversation export
        self.html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f8f9fa;
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            text-align: center;
        }
        .conversation {
            background: white;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .message {
            margin-bottom: 20px;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid;
        }
        .user-message {
            background-color: #e3f2fd;
            border-left-color: #2196f3;
        }
        .assistant-message {
            background-color: #f3e5f5;
            border-left-color: #9c27b0;
        }
        .system-message {
            background-color: #fff3e0;
            border-left-color: #ff9800;
        }
        .message-header {
            display: flex;
            justify-content: between;
            align-items: center;
            margin-bottom: 10px;
        }
        .role {
            font-weight: bold;
            text-transform: capitalize;
        }
        .timestamp {
            font-size: 0.9em;
            color: #666;
            margin-left: auto;
        }
        .knowledge-update {
            background-color: #e8f5e8;
            border-left-color: #4caf50;
            border: 1px solid #4caf50;
        }
        .metadata {
            background-color: #f5f5f5;
            padding: 15px;
            border-radius: 8px;
            margin-top: 20px;
            font-size: 0.9em;
        }
        .footer {
            text-align: center;
            color: #666;
            margin-top: 30px;
            padding: 20px;
            border-top: 1px solid #ddd;
        }
        pre {
            background-color: #f4f4f4;
            padding: 10px;
            border-radius: 4px;
            overflow-x: auto;
        }
        code {
            background-color: #f4f4f4;
            padding: 2px 4px;
            border-radius: 2px;
            font-family: 'Monaco', 'Menlo', monospace;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ title }}</h1>
        {% if subtitle %}<p>{{ subtitle }}</p>{% endif %}
        {% if custom_header %}<p>{{ custom_header }}</p>{% endif %}
    </div>
    
    {% if metadata and include_metadata %}
    <div class="metadata">
        <h3>📊 Conversation Metadata</h3>
        <p><strong>Conversation ID:</strong> {{ metadata.conversation_id }}</p>
        <p><strong>Total Messages:</strong> {{ metadata.message_count }}</p>
        <p><strong>Start Time:</strong> {{ metadata.start_time }}</p>
        <p><strong>Last Activity:</strong> {{ metadata.last_activity }}</p>
        {% if metadata.knowledge_updates > 0 %}
        <p><strong>Knowledge Updates:</strong> {{ metadata.knowledge_updates }}</p>
        {% endif %}
        <p><strong>Export Date:</strong> {{ export_date }}</p>
    </div>
    {% endif %}
    
    <div class="conversation">
        {% for message in messages %}
        <div class="message {% if message.role == 'user' %}user-message{% elif message.role == 'assistant' %}assistant-message{% else %}system-message{% endif %}{% if message.is_knowledge_update %} knowledge-update{% endif %}">
            <div class="message-header">
                <span class="role">
                    {% if message.role == 'user' %}🧑 User{% elif message.role == 'assistant' %}🤖 Assistant{% else %}⚙️ System{% endif %}
                    {% if message.is_knowledge_update %}📚 (Knowledge Update){% endif %}
                </span>
                {% if include_timestamps %}
                <span class="timestamp">{{ message.timestamp }}</span>
                {% endif %}
            </div>
            <div class="content">{{ message.content | safe }}</div>
        </div>
        {% endfor %}
    </div>
    
    {% if custom_footer %}
    <div class="footer">
        {{ custom_footer }}
    </div>
    {% endif %}
    
    <div class="footer">
        <p>Exported from <strong>ATLAS Enterprise</strong> - AI-Powered Trade Compliance Platform</p>
        <p>Generated on {{ export_date }}</p>
    </div>
</body>
</html>
        """
    
    async def initialize(self):
        """Initialize the export service."""
        if self._initialized:
            return
        
        self.cache = get_cache()
        
        # Create export directories
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        
        self._initialized = True
        logger.info("Conversation export service initialized")
    
    async def export_conversation(self, conversation_id: str, options: ExportOptions,
                                user_id: Optional[str] = None) -> ExportResult:
        """Export a conversation in the specified format."""
        try:
            # Get conversation data
            conversation_data = await self._get_conversation_data(conversation_id)
            if not conversation_data:
                return ExportResult(
                    success=False,
                    error=f"Conversation {conversation_id} not found"
                )
            
            # Generate export ID
            export_id = str(uuid.uuid4())
            
            # Apply anonymization if requested
            if options.anonymize_user_data:
                conversation_data = self._anonymize_conversation(conversation_data)
            
            # Filter messages based on options
            filtered_messages = self._filter_messages(conversation_data["messages"], options)
            
            # Generate export based on format
            if options.format == ExportFormat.JSON:
                result = await self._export_to_json(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.PDF:
                result = await self._export_to_pdf(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.MARKDOWN:
                result = await self._export_to_markdown(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.HTML:
                result = await self._export_to_html(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.EXCEL:
                result = await self._export_to_excel(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.TXT:
                result = await self._export_to_txt(conversation_data, filtered_messages, options, export_id)
            elif options.format == ExportFormat.CSV:
                result = await self._export_to_csv(conversation_data, filtered_messages, options, export_id)
            else:
                return ExportResult(
                    success=False,
                    error=f"Unsupported export format: {options.format.value}"
                )
            
            # Log export event
            if result.success and user_id:
                await log_business_event(
                    "conversation_export",
                    {
                        "user_id": user_id,
                        "conversation_id": conversation_id,
                        "export_id": export_id,
                        "format": options.format.value,
                        "template": options.template.value,
                        "file_size": result.file_size
                    }
                )
            
            return result
            
        except Exception as e:
            logger.error(f"Conversation export failed: {e}")
            return ExportResult(
                success=False,
                error=str(e)
            )
    
    async def _get_conversation_data(self, conversation_id: str) -> Optional[Dict[str, Any]]:
        """Get conversation data with messages and metadata."""
        try:
            # Get conversation summary
            summary = await enhanced_ai_service.get_conversation_summary(conversation_id)
            if "error" in summary:
                return None
            
            # Get full conversation history
            messages = await enhanced_ai_service._get_conversation_history(conversation_id, limit=1000)
            
            return {
                "conversation_id": conversation_id,
                "summary": summary,
                "messages": messages,
                "export_timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Failed to get conversation data: {e}")
            return None
    
    def _filter_messages(self, messages: List[ConversationMessage], options: ExportOptions) -> List[Dict[str, Any]]:
        """Filter and format messages based on export options."""
        filtered = []
        
        for message in messages:
            # Skip system messages if not requested
            if message.role == ConversationRole.SYSTEM and not options.include_system_messages:
                continue
            
            # Check for knowledge updates
            is_knowledge_update = "knowledge" in message.content.lower() and "added" in message.content.lower()
            
            # Skip knowledge updates if not requested
            if is_knowledge_update and not options.include_knowledge_updates:
                continue
            
            formatted_message = {
                "role": message.role.value,
                "content": message.content,
                "timestamp": message.timestamp.strftime("%Y-%m-%d %H:%M:%S") if options.include_timestamps else None,
                "is_knowledge_update": is_knowledge_update,
                "metadata": message.metadata or {}
            }
            
            filtered.append(formatted_message)
        
        return filtered
    
    def _anonymize_conversation(self, conversation_data: Dict[str, Any]) -> Dict[str, Any]:
        """Anonymize user data in conversation."""
        # This would implement proper anonymization logic
        # For demo purposes, we'll just replace certain patterns
        
        import re
        
        anonymized = conversation_data.copy()
        
        # Anonymize email patterns
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        
        for message in anonymized.get("messages", []):
            if hasattr(message, 'content'):
                message.content = re.sub(email_pattern, '[EMAIL_REDACTED]', message.content)
        
        return anonymized
    
    async def _export_to_json(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                            options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to JSON format."""
        try:
            export_data = {
                "export_id": export_id,
                "export_date": datetime.now().isoformat(),
                "format": "json",
                "conversation": {
                    "id": conversation_data["conversation_id"],
                    "summary": conversation_data["summary"] if options.include_metadata else None,
                    "messages": messages,
                    "message_count": len(messages)
                },
                "options": {
                    "template": options.template.value,
                    "include_metadata": options.include_metadata,
                    "include_timestamps": options.include_timestamps,
                    "anonymized": options.anonymize_user_data
                }
            }
            
            if options.custom_title:
                export_data["title"] = options.custom_title
            
            # Save to file
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.json"
            file_path = self.export_dir / file_name
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "json", "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_pdf(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                           options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to PDF format."""
        try:
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.pdf"
            file_path = self.export_dir / file_name
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=A4 if options.page_size == "a4" else letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1,  # Center
                textColor=HexColor('#2c3e50')
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=HexColor('#34495e')
            )
            
            user_style = ParagraphStyle(
                'UserMessage',
                parent=styles['Normal'],
                fontSize=options.font_size,
                leftIndent=20,
                rightIndent=20,
                spaceAfter=12,
                backColor=HexColor('#e3f2fd')
            )
            
            assistant_style = ParagraphStyle(
                'AssistantMessage',
                parent=styles['Normal'],
                fontSize=options.font_size,
                leftIndent=20,
                rightIndent=20,
                spaceAfter=12,
                backColor=HexColor('#f3e5f5')
            )
            
            # Build PDF content
            story = []
            
            # Title
            title = options.custom_title or f"AI Conversation Export"
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 20))
            
            # Metadata
            if options.include_metadata:
                summary = conversation_data.get("summary", {})
                metadata_data = [
                    ["Conversation ID", conversation_data["conversation_id"]],
                    ["Total Messages", str(len(messages))],
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ]
                
                if summary.get("start_time"):
                    metadata_data.append(["Start Time", summary["start_time"]])
                if summary.get("last_activity"):
                    metadata_data.append(["Last Activity", summary["last_activity"]])
                
                metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
                metadata_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), HexColor('#f8f9fa')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), HexColor('#2c3e50')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, HexColor('#dee2e6'))
                ]))
                
                story.append(Paragraph("📊 Conversation Metadata", heading_style))
                story.append(metadata_table)
                story.append(Spacer(1, 20))
            
            # Messages
            story.append(Paragraph("💬 Conversation Messages", heading_style))
            story.append(Spacer(1, 12))
            
            for message in messages:
                # Message header
                role_icon = "🧑" if message["role"] == "user" else "🤖" if message["role"] == "assistant" else "⚙️"
                role_text = f"{role_icon} {message['role'].title()}"
                
                if message.get("is_knowledge_update"):
                    role_text += " 📚 (Knowledge Update)"
                
                if options.include_timestamps and message.get("timestamp"):
                    role_text += f" - {message['timestamp']}"
                
                story.append(Paragraph(f"<b>{role_text}</b>", styles['Normal']))
                
                # Message content
                content = message["content"].replace('\n', '<br/>')
                style = user_style if message["role"] == "user" else assistant_style
                story.append(Paragraph(content, style))
                story.append(Spacer(1, 8))
            
            # Custom footer
            if options.custom_footer:
                story.append(Spacer(1, 20))
                story.append(Paragraph(options.custom_footer, styles['Normal']))
            
            # Build PDF
            doc.build(story)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "pdf", "pages": "auto", "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_html(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                            options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to HTML format."""
        try:
            # Prepare template data
            template = Template(self.html_template)
            
            title = options.custom_title or "AI Conversation Export"
            subtitle = f"Conversation ID: {conversation_data['conversation_id']}"
            
            # Convert markdown in messages to HTML
            for message in messages:
                if message["role"] == "assistant":
                    # Convert markdown to HTML for assistant messages
                    message["content"] = markdown.markdown(
                        message["content"],
                        extensions=['codehilite', 'fenced_code']
                    )
            
            html_content = template.render(
                title=title,
                subtitle=subtitle,
                custom_header=options.custom_header,
                custom_footer=options.custom_footer,
                messages=messages,
                metadata=conversation_data.get("summary", {}),
                include_metadata=options.include_metadata,
                include_timestamps=options.include_timestamps,
                export_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            # Save to file
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.html"
            file_path = self.export_dir / file_name
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "html", "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"HTML export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_markdown(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                                options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to Markdown format."""
        try:
            markdown_content = []
            
            # Title
            title = options.custom_title or "AI Conversation Export"
            markdown_content.append(f"# {title}\n")
            
            # Custom header
            if options.custom_header:
                markdown_content.append(f"{options.custom_header}\n")
            
            # Metadata
            if options.include_metadata:
                summary = conversation_data.get("summary", {})
                markdown_content.extend([
                    "## 📊 Conversation Metadata\n",
                    f"- **Conversation ID:** `{conversation_data['conversation_id']}`",
                    f"- **Total Messages:** {len(messages)}",
                    f"- **Export Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ])
                
                if summary.get("start_time"):
                    markdown_content.append(f"- **Start Time:** {summary['start_time']}")
                if summary.get("last_activity"):
                    markdown_content.append(f"- **Last Activity:** {summary['last_activity']}")
                if summary.get("knowledge_updates", 0) > 0:
                    markdown_content.append(f"- **Knowledge Updates:** {summary['knowledge_updates']}")
                
                markdown_content.append("")
            
            # Messages
            markdown_content.append("## 💬 Conversation\n")
            
            for message in messages:
                role_icon = "🧑" if message["role"] == "user" else "🤖" if message["role"] == "assistant" else "⚙️"
                role_text = f"{role_icon} **{message['role'].title()}**"
                
                if message.get("is_knowledge_update"):
                    role_text += " 📚 *(Knowledge Update)*"
                
                if options.include_timestamps and message.get("timestamp"):
                    role_text += f" - *{message['timestamp']}*"
                
                markdown_content.extend([
                    f"### {role_text}\n",
                    f"{message['content']}\n",
                    "---\n"
                ])
            
            # Custom footer
            if options.custom_footer:
                markdown_content.extend([
                    "## Footer\n",
                    f"{options.custom_footer}\n"
                ])
            
            # Add export info
            markdown_content.extend([
                "---",
                "*Exported from **ATLAS Enterprise** - AI-Powered Trade Compliance Platform*",
                f"*Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*"
            ])
            
            # Save to file
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.md"
            file_path = self.export_dir / file_name
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(markdown_content))
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "markdown", "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"Markdown export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_excel(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                             options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to Excel format."""
        try:
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.xlsx"
            file_path = self.export_dir / file_name
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Metadata sheet
            if options.include_metadata:
                ws_meta = wb.active
                ws_meta.title = "Metadata"
                
                # Headers
                ws_meta['A1'] = "Property"
                ws_meta['B1'] = "Value"
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws_meta[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Metadata rows
                summary = conversation_data.get("summary", {})
                metadata_rows = [
                    ["Conversation ID", conversation_data["conversation_id"]],
                    ["Total Messages", len(messages)],
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ]
                
                if summary.get("start_time"):
                    metadata_rows.append(["Start Time", summary["start_time"]])
                if summary.get("last_activity"):
                    metadata_rows.append(["Last Activity", summary["last_activity"]])
                if summary.get("knowledge_updates", 0) > 0:
                    metadata_rows.append(["Knowledge Updates", summary["knowledge_updates"]])
                
                for i, (prop, value) in enumerate(metadata_rows, 2):
                    ws_meta[f'A{i}'] = prop
                    ws_meta[f'B{i}'] = value
                
                # Auto-size columns
                for column in ws_meta.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_meta.column_dimensions[column_letter].width = adjusted_width
                
                # Messages sheet
                ws_messages = wb.create_sheet("Messages")
            else:
                ws_messages = wb.active
                ws_messages.title = "Messages"
            
            # Message headers
            headers = ["#", "Role", "Content"]
            if options.include_timestamps:
                headers.append("Timestamp")
            headers.extend(["Knowledge Update", "Message Length"])
            
            for i, header in enumerate(headers, 1):
                cell = ws_messages.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Message data
            for i, message in enumerate(messages, 2):
                ws_messages.cell(row=i, column=1, value=i-1)  # Message number
                ws_messages.cell(row=i, column=2, value=message["role"].title())
                ws_messages.cell(row=i, column=3, value=message["content"])
                
                col = 4
                if options.include_timestamps:
                    ws_messages.cell(row=i, column=col, value=message.get("timestamp", ""))
                    col += 1
                
                ws_messages.cell(row=i, column=col, value="Yes" if message.get("is_knowledge_update") else "No")
                ws_messages.cell(row=i, column=col+1, value=len(message["content"]))
                
                # Color-code messages
                if message["role"] == "user":
                    fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                elif message["role"] == "assistant":
                    fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
                else:
                    fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                
                for cell in ws_messages[i]:
                    cell.fill = fill
            
            # Auto-size columns
            for column in ws_messages.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # Limit content column width
                if column_letter == 'C':  # Content column
                    adjusted_width = min(max_length + 2, 80)
                else:
                    adjusted_width = min(max_length + 2, 30)
                ws_messages.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "excel", "sheets": len(wb.worksheets), "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_txt(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                           options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to plain text format."""
        try:
            content_lines = []
            
            # Title
            title = options.custom_title or "AI Conversation Export"
            content_lines.extend([
                "=" * len(title),
                title,
                "=" * len(title),
                ""
            ])
            
            # Custom header
            if options.custom_header:
                content_lines.extend([options.custom_header, ""])
            
            # Metadata
            if options.include_metadata:
                summary = conversation_data.get("summary", {})
                content_lines.extend([
                    "CONVERSATION METADATA",
                    "-" * 20,
                    f"Conversation ID: {conversation_data['conversation_id']}",
                    f"Total Messages: {len(messages)}",
                    f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ])
                
                if summary.get("start_time"):
                    content_lines.append(f"Start Time: {summary['start_time']}")
                if summary.get("last_activity"):
                    content_lines.append(f"Last Activity: {summary['last_activity']}")
                if summary.get("knowledge_updates", 0) > 0:
                    content_lines.append(f"Knowledge Updates: {summary['knowledge_updates']}")
                
                content_lines.extend(["", ""])
            
            # Messages
            content_lines.extend([
                "CONVERSATION MESSAGES",
                "-" * 22,
                ""
            ])
            
            for i, message in enumerate(messages, 1):
                role_text = message["role"].upper()
                if message.get("is_knowledge_update"):
                    role_text += " (KNOWLEDGE UPDATE)"
                
                content_lines.append(f"[{i}] {role_text}")
                
                if options.include_timestamps and message.get("timestamp"):
                    content_lines.append(f"Time: {message['timestamp']}")
                
                content_lines.extend([
                    "-" * 40,
                    message["content"],
                    "",
                    ""
                ])
            
            # Custom footer
            if options.custom_footer:
                content_lines.extend([
                    "FOOTER",
                    "-" * 6,
                    options.custom_footer,
                    ""
                ])
            
            # Export info
            content_lines.extend([
                "=" * 50,
                "Exported from ATLAS Enterprise",
                "AI-Powered Trade Compliance Platform",
                f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                "=" * 50
            ])
            
            # Save to file
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.txt"
            file_path = self.export_dir / file_name
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(content_lines))
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "txt", "lines": len(content_lines), "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"TXT export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def _export_to_csv(self, conversation_data: Dict[str, Any], messages: List[Dict[str, Any]], 
                           options: ExportOptions, export_id: str) -> ExportResult:
        """Export conversation to CSV format."""
        try:
            import csv
            
            file_name = f"conversation_{conversation_data['conversation_id']}_{export_id}.csv"
            file_path = self.export_dir / file_name
            
            # Define CSV headers
            headers = ["message_number", "role", "content", "message_length", "knowledge_update"]
            if options.include_timestamps:
                headers.append("timestamp")
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                writer.writerow(headers)
                
                # Write message data
                for i, message in enumerate(messages, 1):
                    row = [
                        i,
                        message["role"],
                        message["content"].replace('\n', ' ').replace('\r', ' '),  # Clean newlines
                        len(message["content"]),
                        "Yes" if message.get("is_knowledge_update") else "No"
                    ]
                    
                    if options.include_timestamps:
                        row.append(message.get("timestamp", ""))
                    
                    writer.writerow(row)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_name,
                file_size=file_size,
                export_id=export_id,
                download_url=f"/api/v1/conversations/download/{export_id}",
                metadata={"format": "csv", "rows": len(messages) + 1, "messages": len(messages)}
            )
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return ExportResult(success=False, error=str(e))
    
    async def get_export_file(self, export_id: str) -> Optional[Path]:
        """Get export file by export ID."""
        try:
            # Find file with this export ID
            for file_path in self.export_dir.glob(f"*_{export_id}.*"):
                if file_path.is_file():
                    return file_path
            return None
            
        except Exception as e:
            logger.error(f"Failed to get export file: {e}")
            return None
    
    async def cleanup_old_exports(self, max_age_hours: int = 24):
        """Clean up old export files."""
        try:
            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
            
            deleted_count = 0
            for file_path in self.export_dir.glob("*"):
                if file_path.is_file():
                    file_modified = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_modified < cutoff_time:
                        file_path.unlink()
                        deleted_count += 1
            
            logger.info(f"Cleaned up {deleted_count} old export files")
            return deleted_count
            
        except Exception as e:
            logger.error(f"Export cleanup failed: {e}")
            return 0


# Global export service instance
conversation_export_service = ConversationExportService() 